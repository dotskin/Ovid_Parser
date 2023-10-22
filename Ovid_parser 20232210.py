# This program is useful in meta-analysis research project, as it can neatly 
# take all the informations from a list of references, after the screening 
# from title and abstract has been done, and save them in a excel workbook.
# In order to use employ it, you need to download the full list of references 
# from Ovid MEDLINE in "Ovid labeled Citation" format and apply a three dollars 
# symbol ($$$) after the reference number of an article you want to include. 
# Convert the file into .txt type, and change its name to match that in line 92 
# of this program. Finally, execute this program in the same directory of 
# the input file.
# IMPORTANT: In order for this program to work, the input file downloaded
# from Ovid MEDLINE needs to be in "Ovid Labeled Citation" format.

import re
import openpyxl

# Function that extract the list of lines between a line that starts with
# '$$$' and a line that ends with 'SFX'. Each list of lines is a block of
# lines that will populate the list called blocks, which is returned by
# the function. 
def extract_blocks(filename):
    blocks = []
    current_block = []
    in_block = False
    with open(filename) as f:
        for line in f:
            if "$$$" in line:
                in_block = True
                current_block = []
                current_block.append(line)
            
            elif "SFX" in line:
                if current_block:
                    blocks.append(current_block)
                    current_block = []
                    in_block = False

            else:
                if in_block:
                    current_block.append(line)

    if current_block:
        blocks.append(current_block)
    return blocks

# Function that extract list of lines at or after a pattern is matched
# in a line of the blocks of text. In order for the patterns to work, the
# input file taken from Ovid needs to be in "Ovid Labeled Citation" format.
def extract_infos(pattern, blocks):
    # Initialize a list to store the location of the years of publication
    info_indexes = []
    for block in blocks:
        found_match = False
    
        # Starts a counter for the index of the line that matches our search
        # with one special rules for one of the patterns
        if pattern == re.compile(r'\A\<'):
            n = -1
        else:
            n = 0        
        for line in block:
            n = n + 1
        
            # Search for the year of publication in each block
            match = pattern.search(line)
        
            # If a match is found, extract the location and add it to the list
            if match:
                info_indexes.append(n)
                found_match = True
            
            # This part is here in order to always make sure to add at least
            # one element to the info_indexes, even if the pattern is not
            # matched. Otherwise it would misalign all the dataframe!
            if line == block[-1] and found_match == False:
                info_indexes.append(1)

        # Initialize a list to store the matched informations
        info_list = []

        # Create a counter that runs along the list of indexes
        n = -1

    # Run through each block, and append every line that appear at the index
    # location to a list with relevant lines from each block
    for block in blocks:
        n = n + 1
        info_list.append((block[int(info_indexes[n])]).strip())
    return info_list

# MAIN CODE STARTS HERE

blocks = extract_blocks("input_example.txt")

# The patterns we will use to search for the year of publication
pattern = re.compile(r'\AYear of Publication\n')
years = extract_infos(pattern, blocks)
pattern = re.compile(r'Title\n')
titles = extract_infos(pattern, blocks)
pattern = re.compile(r'\AAuthor(s)?\n')
authors = extract_infos(pattern, blocks)
pattern = re.compile(r'\A\<')
nsearches = extract_infos(pattern, blocks)

# Last lines will check if there is any missing information from articles
print (f'In total {len(blocks)} articles have passed the screening\n')
if (len(years) or len(titles) or len(authors) or len(nsearches)) != len(blocks):
    print('ERROR! The lists with all the information do not match!!!')
elif (len(years) or len(titles) or len(authors) or len(nsearches)) == len(blocks):
    print('The number of articles matches with other parameters!\n')

# Create a new workbook object
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

worksheet.cell(row=1, column=1, value='N SEARCH')
worksheet.cell(row=1, column=2, value='AUTHORS')
worksheet.cell(row=1, column=3, value='YEARS')
worksheet.cell(row=1, column=4, value='TITLES')

# Loop through the data and write it to the worksheet.
for i in range(len(nsearches)):
    worksheet.cell(row=i+2, column=1, value=str(nsearches[i]))
    worksheet.cell(row=i+2, column=2, value=authors[i])
    worksheet.cell(row=i+2, column=3, value=years[i])
    worksheet.cell(row=i+2, column=4, value=titles[i])

# Save the workbook in the ordered format.
workbook.save('output.xlsx')